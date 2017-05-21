from openpyxl import load_workbook
import numpy as np
import pylab as pl
from amigo import geom
import scipy as sp
from collections import OrderedDict
from itertools import cycle, count
from scipy.linalg import norm
from scipy.optimize import minimize
from scipy.interpolate import InterpolatedUnivariateSpline as IUS
from scipy.signal import savgol_filter
import nova
from amigo.IO import class_dir
import seaborn as sns
rc = {'figure.figsize': [7, 7 * 12 / 9], 'savefig.dpi': 100,
      'savefig.jpeg_quality': 100, 'savefig.pad_inches': 0.1,
      'lines.linewidth': 1.5}
sns.set(context='talk', style='white', font='sans-serif', palette='Set2',
        font_scale=7 / 8, rc=rc)
color = sns.color_palette('Paired', 12)
ic = count(0)

nova_root = class_dir(nova)


def get_label(label, label_array, force=False, part=None):
    if label is not None:
        label_array.append(label.replace(' ', '_'))
        flag = True
    else:
        if force:
            label_array.append(part)
            if len(label_array) == 1:
                flag = True
            else:
                if label_array[-2] != label_array[-1]:
                    flag = True
                else:
                    flag = False
        else:
            flag = False
    return flag


def add_segment(lines, k):
    segment = '{:1.0f}'.format(next(k))
    lines[segment] = {'x': np.array([]), 'z': np.array([])}
    return segment


def append_value(lines, segment, x, z):
    lines[segment]['x'] = np.append(lines[segment]['x'], x)
    lines[segment]['z'] = np.append(lines[segment]['z'], z)


def cutcorners(x, z):
    x, z = geom.pointloop(x, z, ref='min')  # form loop from closest neighbour
    n = len(x) - 1
    dX = np.array([x[1:] - x[:-1], z[1:] - z[:-1]])
    dX = np.array([np.gradient(x), np.gradient(z)])
    dL = norm(dX, axis=0)
    k, kflag = count(0), False
    lines = OrderedDict()
    segment = add_segment(lines, k)
    for i in range(n):
        dot = np.dot(dX[:, i] / dL[i], dX[:, i + 1] / dL[i + 1])
        dot = dot if dot < 1.0 else 1.0
        kink = np.arccos(dot) * 180 / np.pi
        append_value(lines, segment, x[i + 1], z[i + 1])
        if abs(kink) > 40 and kflag is False:  # angle, deg
            segment = add_segment(lines, k)
            append_value(lines, segment, x[i + 1], z[i + 1])
            kflag = True
        else:
            kflag = False
    segments = list(lines.keys())
    l = np.zeros(len(segments))
    for i, seg in enumerate(segments):
        l[i] = len(lines[seg]['x'])
    seg = np.argsort(l)[-2:]  # select loops (in/out)
    rmax = np.zeros(len(seg))
    for i, s in enumerate(seg):
        rmax[i] = np.max(lines[segments[s]]['x'])
    seg = seg[np.argsort(rmax)]
    lines_sort = OrderedDict()
    for s in seg[:2]:
        lines_sort[segments[s]] = lines[segments[s]]
    return lines_sort


def cluster_points(x, z):
    R, Z = geom.pointloop(x, z)
    dX = norm([x[1:] - x[:-1], z[1:] - z[:-1]], axis=0)
    dx_median = sp.median(dX)
    cluster, i = OrderedDict(), count(0)
    for x, z in zip(R, Z):
        dx = []
        for cl in cluster:
            rc, zc = cluster[cl]['x'], cluster[cl]['z']
            dx.append(np.min(norm([x - rc, z - zc], axis=0)))
        if len(dx) == 0 or np.min(dx) > 2 * dx_median:  # new
            cl = 'group{:1.0f}'.format(next(i))
            cluster[cl] = {}
            cluster[cl] = {'x': [x], 'z': [z]}
        else:
            icl = np.argmin(dx)
            cl = list(cluster.keys())[icl]
            cluster[cl]['x'] = np.append(cluster[cl]['x'], x)
            cluster[cl]['z'] = np.append(cluster[cl]['z'], z)
    for cl in cluster:
        x, z = cluster[cl]['x'], cluster[cl]['z']
        dx = norm([x[:1] - x[:-1], z[:1] - z[:-1]], axis=0)
        imax = np.argmax(dx) + 1
        x = np.append(x[imax:], x[:imax])
        z = np.append(z[imax:], z[:imax])
        cluster[cl]['x'], cluster[cl]['z'] = x, z
    return cluster


def set_figure():
    pl.axis('off')
    pl.axis('equal')


class DEMO(object):

    def __init__(self):
        self.filename = 'DEMO1_Reference_Design_-_2015_April_(_EU_2MG46D_v1_0'
        # self.filename = 'SN-A31-k165'
        self.read(self.filename)
        self.process()
        self.get_limiters()
        self.get_ports()
        self.get_fw()

    def read(self, filename):
        ref = nova_root + '/../Data/'
        # read_only=True,data_only=True
        wb = load_workbook(filename=ref + filename + '.xlsx')
        ws = wb[wb.get_sheet_names()[0]]

        self.parts = OrderedDict()  # component parts
        part, loop = [], []
        for row in ws.columns:
            new_part = get_label(row[0].value, part)
            if new_part:
                self.parts[part[-1]] = OrderedDict()
            if len(part) == 0:
                continue
            new_loop = get_label(row[1].value, loop,
                                 force=new_part, part=part[-1])
            p, l = part[-1], loop[-1]
            if new_loop:
                self.parts[p][l] = OrderedDict()

            comp, start = row[2].value, 2
            while comp is None:
                start += 1
                comp = row[start].value
            if comp == 'Rad':  # replace confusing 'Rad' label
                comp = 'x'

            self.parts[p][l][comp] = np.zeros(len(row) - 1)
            for i, x in enumerate(row[start + 1:]):
                try:
                    self.parts[p][l][comp][i] = 1e-3 * float(x.value)  # m
                except:
                    break
            self.parts[p][l][comp] = self.parts[p][l][comp][:i]

    def process(self):
        for part in self.parts:
            p = {'in': {'x': np.array([]), 'z': np.array([])},
                 'out': {'x': np.array([]), 'z': np.array([])},
                 'ports': {'x': np.array([]), 'z': np.array([])},
                 'x': np.array([]), 'z': np.array([])}

            for loop, side in zip(self.parts[part], ['out', 'in', 'ports']):
                x, z = geom.read_loop(self.parts[part], loop)
                p[side]['x'], p[side]['z'] = x, z

            if part in ['TF_Coil', 'Vessel', 'Blanket'] and len(x) > 0:
                if side != 'out':
                    p['x'], p['z'] = geom.polyloop(p['in'], p['out'])
                else:
                    p['x'], p['z'] = geom.pointloop(
                        p['out']['x'], p['out']['z'])
                    lines = cutcorners(p['x'], p['z'])  # select halfs
                    for seg, side in zip(lines, ['in', 'out']):
                        p[side] = {'x': lines[seg]['x'], 'z': lines[seg]['z']}
            for key in p:
                self.parts[part][key] = p[key]

    def get_ports(self, plot=False):
        points = self.parts['Vessel']['ports']
        clusters = cluster_points(points['x'], points['z'])
        port = OrderedDict()
        for i, cl in enumerate(clusters):
            x, z = clusters[cl]['x'], clusters[cl]['z']
            switch = x.max() - x.min() < 0.5 * (z.max() - z.min())
            if switch:  # rotate coordinates
                r, y = np.copy(z), np.copy(x)
            else:
                r, y = np.copy(x), np.copy(z)
            index = np.argsort(r)
            r, y, x, z = r[index], y[index], x[index], z[index]
            M = np.array([np.ones(len(r)), r]).T  # linear least-squares fit
            a = sp.linalg.lstsq(M, y)[0]
            fit = a[0] + a[1] * r
            if switch:
                x_fit, z_fit = fit, z
            else:
                x_fit, z_fit = x, fit
            n_hat = np.array([-(z_fit[-1] - z_fit[0]),
                              x_fit[-1] - x_fit[0]])
            n_hat /= norm(n_hat)
            n = len(x)
            count = {'left': 0, 'right': 0}
            p = 'P{:1.0f}'.format(i)
            port[p] = {'left': {'x': np.zeros(n), 'z': np.zeros(n)},
                       'right': {'x': np.zeros(n), 'z': np.zeros(n)}}
            for x_, z_ in zip(x, z):
                for dot, side in zip([1, -1], ['left', 'right']):
                    if dot * np.dot([x_ - x_fit[0], z_ - z_fit[0]], n_hat) > 0:
                        port[p][side]['x'][count[side]] = x_
                        port[p][side]['z'][count[side]] = z_
                        count[side] += 1
            for side in ['left', 'right']:
                for var in ['x', 'z']:  # trim
                    n = count[side]
                    port[p][side][var] = port[p][side][var][:n]
        ro = np.mean(self.parts['Blanket']['x'])
        zo = np.mean(self.parts['Blanket']['z'])
        theta = np.zeros(len(port))
        for i, p in enumerate(port):
            theta[i] = np.arctan2(port[p]['left']['z'][0] - zo,
                                  port[p]['left']['x'][0] - ro)
        index = list(np.argsort(theta))
        pkey = list(port.keys())
        self.port = OrderedDict()
        for i, indx in enumerate(index):
            psort = 'P{:1.0f}'.format(i)
            self.port[psort] = port[pkey[indx]]

        if plot:
            self.plot_ports()

    def plot_ports(self):
        for p in self.port:
            for s in self.port[p]:
                if s == 'left':
                    c = color[8]
                else:
                    c = color[8]
                pl.plot(self.port[p][s]['x'], self.port[p][s]['z'],
                        zorder=3, color=c, lw=1)

    def get_limiters(self, plot=False):
        x = self.parts['Plasma']['out']
        self.limiter = OrderedDict()
        clusters = cluster_points(x['x'], x['z'])
        for i, cl in enumerate(clusters):
            x, z = clusters[cl]['x'], clusters[cl]['z']
            self.limiter['L{:1.0f}'.format(i)] = {'x': x, 'z': z}
            if plot:
                pl.plot(x, z, color=0.5 * np.ones(3))

    def plot_limiter(self):
        pl.plot(self.limiter['L3']['x'],
                self.limiter['L3']['z'], color=0.6 * np.ones(3))

    def blanket_thickness(self, Nt=100, plot=False):
        bl, loop = {}, {}  # read blanket
        for side in ['in', 'out']:
            bl[side], c = {}, {}
            for x in ['x', 'z']:
                c[x] = self.parts['Blanket'][side][x]
            x, z = geom.order(c['x'], c['z'], anti=True)
            x, z, l = geom.unique(x, z)  # remove repeats
            bl[side]['x'], bl[side]['z'] = x, z
            loop[side] = {'x': IUS(l, x), 'z': IUS(l, z)}  # interpolator

        def thickness(L, Lo, loop, norm):
            x = loop['in']['x'](Lo) + L[0] * norm['nr'](Lo)
            z = loop['in']['z'](Lo) + L[0] * norm['nz'](Lo)
            err = (x - loop['out']['x'](L[1]))**2 +\
                  (z - loop['out']['z'](L[1]))**2
            return err

        x, z = geom.unique(bl['in']['x'], bl['in']['z'])[:2]  # remove repeats
        nr, nz, x, z = geom.normal(x, z)
        l = geom.length(x, z)
        norm = {'nr': IUS(l, nr), 'nz': IUS(l, nz)}  # interpolator

        dt, Lo = np.zeros(Nt), np.linspace(0, 1, Nt)
        for i, lo in enumerate(Lo):
            L = minimize(thickness, [0, lo], method='L-BFGS-B',
                         bounds=([0, 5], [0, 1]), args=(lo, loop, norm)).x
            dt[i] = np.sqrt((loop['in']['x'](lo) - loop['out']['x'](L[1]))**2 +
                            (loop['in']['z'](lo) - loop['out']['z'](L[1]))**2)
            dt[i] = L[0]
        dt = savgol_filter(dt, 7, 2)  # filter
        if plot:
            pl.plot(Lo, dt)
        return [np.min(dt), np.max(dt)]

    def get_fw(self, plot=False):
        rbl = self.parts['Blanket']['in']['x']  # read blanket
        zbl = self.parts['Blanket']['in']['z']
        zmin = np.zeros(len(self.limiter))  # select divertor limiter
        for i, limiter in enumerate(self.limiter):
            zmin[i] = np.min(self.limiter[limiter]['z'])
        imin = np.argmin(zmin)
        div = list(self.limiter.keys())[imin]
        rdiv = self.limiter[div]['x']  # set divertor profile
        zdiv = self.limiter[div]['z']
        cut = np.zeros(2, dtype=int)  # cut and join divertor
        for i, j in enumerate([0, -1]):
            cut[i] = np.argmin(norm([rbl[j] - rdiv, zbl[j] - zdiv], axis=0))
        cut = np.sort(cut)
        x = np.append(rbl, rdiv[cut[0]:cut[1]])
        z = np.append(zbl, zdiv[cut[0]:cut[1]])
        x, z = np.append(x, x[0]), np.append(z, z[0])
        x, z = geom.rzSLine(x, z)
        x, z = x[::-1], z[::-1]  # flip
        self.fw = {'x': x, 'z': z}
        if plot:
            pl.plot(x, z, color=0.5 * np.ones(3))

    def fill_loops(self):
        for part in self.parts:
            geom.polyfill(self.parts[part]['x'],
                          self.parts[part]['z'], color=color[next(ic)])
        set_figure()

    def fill_part(self, part, alpha=1, **kwargs):
        if part == 'TF_Coil':
            cindex = len(self.parts.keys()) - 1
        else:
            cindex = list(self.parts.keys()).index(part) - 1
        c = kwargs.get('color', color[cindex])
        geom.polyfill(self.parts[part]['x'], self.parts[part]['z'],
                      color=c, alpha=alpha)

    def plot(self):
        for part in self.parts:
            for loop in self.parts[part]:
                try:
                    pl.plot(self.parts[part][loop]['x'],
                            self.parts[part][loop]['z'], '.', markersize=5.0)
                except:
                    pass

    def write(self):
        filename = 'DEMO1_sorted'
        wb = load_workbook(filename='./referance/' + self.filename + '.xlsx')

        ws = wb[wb.get_sheet_names()[0]]

        part = iter(self.parts)
        component = cycle(['x', 'z'])
        for row in ws.columns:
            if row[0].value is not None:
                pt = next(part)
                row[0].value = pt
                loop = iter(['out', 'in', 'ports'])
            if row[1].value is not None:
                lp = next(loop)
                row[1].value = lp
            if row[2].value is not None:
                cp = next(component)
                row[2].value = cp
            '''
            for i,x in enumerate(row[3:]):
                if x.value is not None:
                    print(pt,lp,cp,i)
                    x.value = self.parts[pt][lp][cp][i]
            '''

            '''
            new_loop = get_label(row[1].value,loop,
                                 force=new_part,part=part[-1])
            p,l = part[-1],loop[-1]
            if new_loop:
                self.parts[p][l] = OrderedDict()

            comp,start = row[2].value,2
            while comp == None:
                start += 1
                comp = row[start].value
            if comp == 'Rad':  # replace confusing 'Rad' label
                comp = 'x'

            self.parts[p][l][comp] = np.zeros(len(row)-1)
            for i,x in enumerate(row[start+1:]):
                try:
                    self.parts[p][l][comp][i] = 1e-3*float(x.value)  # m
                except:
                    break
            self.parts[p][l][comp] = self.parts[p][l][comp][:i]
            '''
        wb.save('./referance/' + filename + '.xlsx')


if __name__ is '__main__':
    demo = DEMO()

    demo.fill_loops()

    demo.plot()
    demo.plot_ports()
    # demo.write()
