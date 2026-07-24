"""Top-level workbook <-> canonical transcoder.

This module is the single transaction boundary between the measured
spreadsheet corpus and its git-safe canonical form. It ingests a workbook
into a :class:`CanonicalUnit` (long-format rows plus a provenance sidecar),
writes that unit to disk, reads it back, and rebuilds an openpyxl workbook
from it. A directory sweep applies the same ingest to every workbook in a
tree, skipping lock stubs and recording -- rather than aborting on -- files
that fail to open.

The canonical bytes are the CSV bytes; a sha256 over them is the unit's
content address. The round trip workbook -> canonical -> workbook preserves
every numeric measurement exactly.
"""

from __future__ import annotations

from dataclasses import dataclass
import os
from pathlib import Path

import openpyxl

from nova.assembly.canonical import longformat, sidecar, workbook
from nova.assembly.canonical.longformat import Row
from nova.assembly.provenance import digest, sourcemeta

# Revision-marking filename prefixes, longest first so the most specific
# match wins. An empty result means a released (non-private) revision.
_REVISION_PREFIXES = ("_Fprime_", "_inner_ccl_", "__", "_")

# Lock stubs Excel leaves beside an open workbook.
_LOCK_PREFIX = "~$"

_HEADERS = ["Coil", "Point", "Name", "X", "Y", "Z", "uX", "uY", "uZ"]
_TRANSFORM_LABELS = [
    "dx [mm]",
    "dy [mm]",
    "dz [mm]",
    "X [deg]",
    "Y [deg]",
    "Z [deg]",
]
_METADATA_SHEET = "Metadata"
# Placeholder written for a formula cell whose cached value is absent, so the
# rebuilt cell re-ingests as a formula rather than an empty cell.
_FORMULA_PLACEHOLDER = "=NA()"


@dataclass
class CanonicalUnit:
    """One workbook's canonical form: rows plus provenance sidecar.

    Parameters
    ----------
    rows : list of longformat.Row
        Long-format measurement rows.
    sidecar : dict
        Provenance and layout sidecar mapping.
    """

    rows: list[Row]
    sidecar: dict

    @property
    def csv_bytes(self) -> bytes:
        """Return the canonical CSV bytes."""
        return longformat.rows_to_csv_bytes(self.rows)

    @property
    def digest(self) -> str:
        """Return the sha256 content address of the canonical CSV."""
        return digest.digest_bytes(self.csv_bytes)

    @property
    def stem(self) -> str:
        """Return the source filename stem used to name output files."""
        return Path(self.sidecar["source"]["filename"]).stem


def _split_prefix(name: str) -> tuple[str, str]:
    """Split a revision prefix from a filename.

    Parameters
    ----------
    name : str
        Filename including any prefix.

    Returns
    -------
    tuple
        ``(prefix, remainder)``; ``prefix`` is empty for a released revision.
    """
    for prefix in _REVISION_PREFIXES:
        if name.startswith(prefix):
            return prefix, name[len(prefix) :]
    return "", name


def _filename_meta(path) -> dict:
    """Parse sector, IDM document, version, and revision prefix from a name.

    Parameters
    ----------
    path : os.PathLike or str
        Source path; only the basename is inspected.

    Returns
    -------
    dict
        Metadata mapping with keys ``sector``, ``idm_doc``, ``version``,
        ``private``, ``revision_prefix``.
    """
    name = Path(path).name
    prefix, remainder = _split_prefix(name)
    meta = sourcemeta.parse_filename(remainder)
    return {
        "sector": meta.sector,
        "idm_doc": meta.idm_doc,
        "version": meta.version,
        "private": prefix != "",
        "revision_prefix": prefix,
    }


def _source_info(path) -> dict:
    """Build the source-identity mapping for a workbook.

    Parameters
    ----------
    path : os.PathLike or str
        Source workbook.

    Returns
    -------
    dict
        Mapping with ``filename``, ``sha256``, ``size``, ``mtime``.
    """
    path = Path(path)
    stat = path.stat()
    return {
        "filename": path.name,
        "sha256": digest.digest_file(path),
        "size": stat.st_size,
        "mtime": digest.mtime_iso(stat.st_mtime),
    }


def ingest_workbook(xlsx_path) -> CanonicalUnit:
    """Ingest a workbook into its canonical unit.

    Parameters
    ----------
    xlsx_path : os.PathLike or str
        Source workbook, opened read-only.

    Returns
    -------
    CanonicalUnit
        Long-format rows and the provenance sidecar.
    """
    parse = workbook.parse_workbook(xlsx_path)
    rows = longformat.rows_from_parse(parse)
    csv_sha = digest.digest_bytes(longformat.rows_to_csv_bytes(rows))
    document = sidecar.build_sidecar(
        parse,
        source=_source_info(xlsx_path),
        filename_meta=_filename_meta(xlsx_path),
        csv_sha256=csv_sha,
    )
    return CanonicalUnit(rows=rows, sidecar=document)


def write_unit(unit: CanonicalUnit, out_dir) -> Path:
    """Write a canonical unit's CSV and sidecar to a directory.

    Parameters
    ----------
    unit : CanonicalUnit
        Unit to write.
    out_dir : os.PathLike or str
        Destination directory, created if absent.

    Returns
    -------
    pathlib.Path
        Path to the written CSV file.
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / f"{unit.stem}.csv"
    with open(csv_path, "wb") as handle:
        handle.write(unit.csv_bytes)
    sidecar.write_sidecar(unit.sidecar, out_dir / f"{unit.stem}.meta.yaml")
    return csv_path


def read_unit(csv_path) -> CanonicalUnit:
    """Read a canonical unit from a CSV path and its sibling sidecar.

    Parameters
    ----------
    csv_path : os.PathLike or str
        Path to the canonical CSV; the sidecar is ``<stem>.meta.yaml``
        beside it.

    Returns
    -------
    CanonicalUnit
        The reconstructed unit.
    """
    csv_path = Path(csv_path)
    with open(csv_path, "rb") as handle:
        rows = longformat.csv_bytes_to_rows(handle.read())
    document = sidecar.load_sidecar(csv_path.parent / f"{csv_path.stem}.meta.yaml")
    return CanonicalUnit(rows=rows, sidecar=document)


def _write_measure(worksheet, row, col, value, is_formula) -> None:
    """Write one measurement cell, re-emitting formulas as such.

    A formula cell whose value is absent is written as a placeholder formula
    so it re-ingests as a formula; a formula with a cached value is resolved
    to that value; a plain value is written as a number; an absent value
    leaves the cell empty.
    """
    if is_formula and value is None:
        worksheet.cell(row, col, _FORMULA_PLACEHOLDER)
    elif value is not None:
        worksheet.cell(row, col, value)


def _chunk_fiducials(rows):
    """Group fiducial rows into per-point triples in order.

    Parameters
    ----------
    rows : list of longformat.Row
        Fiducial rows for one block, in ``x``, ``y``, ``z`` order.

    Yields
    ------
    list of longformat.Row
        Consecutive triples, one per measured point.
    """
    for start in range(0, len(rows), len(workbook.COORD_AXES)):
        yield rows[start : start + len(workbook.COORD_AXES)]


def _rebuild_block(worksheet, block, transform_rows, fiducial_rows) -> None:
    """Rebuild one coil block (and its transform) into a worksheet."""
    header_row = block["header_row"]
    header_col = block["header_col"]
    width = block["width"]
    for offset in range(width):
        worksheet.cell(header_row, header_col + offset, _HEADERS[offset])

    transform = block.get("transform")
    if transform is not None:
        _rebuild_transform(worksheet, header_col, transform, transform_rows)

    row_index = header_row + 1
    wrote_coil = False
    for triple in _chunk_fiducials(fiducial_rows):
        head = triple[0]
        if not wrote_coil and head.coil is not None:
            worksheet.cell(row_index, header_col, head.coil)
            wrote_coil = True
        if head.point_group_raw != "":
            worksheet.cell(row_index, header_col + 1, head.point_group_raw)
        if head.feature_raw != "":
            worksheet.cell(row_index, header_col + 2, head.feature_raw)
        for cell in triple:
            axis = cell.axis
            _write_measure(
                worksheet,
                row_index,
                header_col + workbook._OFF_COORD[axis],
                cell.value,
                cell.is_formula,
            )
            if (
                width >= workbook.WIDTH_WITH_UNCERTAINTY
                and cell.uncertainty is not None
            ):
                worksheet.cell(
                    row_index,
                    header_col + workbook._OFF_UNC[axis],
                    cell.uncertainty,
                )
        row_index += 1


def _rebuild_transform(worksheet, header_col, transform, transform_rows) -> None:
    """Rebuild a transform block: labels always, values when populated."""
    label_row = transform["label_row"]
    worksheet.cell(label_row, header_col + 2, "transform")
    for index, label in enumerate(_TRANSFORM_LABELS):
        worksheet.cell(label_row, header_col + 3 + index, label)
    if not transform["populated"]:
        return
    by_axis = {row.axis: row for row in transform_rows}
    for axis in workbook.TRANSFORM_AXES:
        row = by_axis.get(axis)
        if row is None:
            continue
        _write_measure(
            worksheet,
            label_row + 1,
            header_col + workbook._OFF_TRANSFORM[axis],
            row.value,
            row.is_formula,
        )


def egress_workbook(unit: CanonicalUnit):
    """Rebuild an openpyxl workbook from a canonical unit.

    The rebuilt workbook places headers, transforms, and values at the
    positions recorded in the sidecar, so that re-ingesting it yields the
    same canonical bytes.

    Parameters
    ----------
    unit : CanonicalUnit
        Unit to rebuild.

    Returns
    -------
    openpyxl.Workbook
        The rebuilt workbook.
    """
    book = openpyxl.Workbook()
    book.remove(book.active)

    metadata_text = unit.sidecar.get("metadata_text") or []
    if metadata_text:
        sheet = book.create_sheet(_METADATA_SHEET)
        for entry in metadata_text:
            sheet.cell(entry["row"], entry["col"], entry["text"])

    rows_by_phase: dict[str, list[Row]] = {}
    for row in unit.rows:
        rows_by_phase.setdefault(row.phase, []).append(row)

    for sheet_layout in unit.sidecar["sheets"]:
        name = sheet_layout["name"]
        worksheet = book.create_sheet(name)
        phase_rows = rows_by_phase.get(name, [])
        for block in sheet_layout["blocks"]:
            coil = block["coil"]
            transform_rows = [
                row
                for row in phase_rows
                if row.record_kind == longformat.RECORD_TRANSFORM and row.coil == coil
            ]
            fiducial_rows = [
                row
                for row in phase_rows
                if row.record_kind == longformat.RECORD_FIDUCIAL and row.coil == coil
            ]
            _rebuild_block(worksheet, block, transform_rows, fiducial_rows)
    return book


def ingest_tree(src_dir, out_dir) -> dict:
    """Ingest every workbook in a directory into canonical units.

    Lock stubs (names beginning ``~$``) are skipped by name. A workbook that
    fails to open is recorded and the sweep continues.

    Parameters
    ----------
    src_dir : os.PathLike or str
        Directory to sweep for ``*.xlsx`` files (non-recursive).
    out_dir : os.PathLike or str
        Destination directory for canonical units.

    Returns
    -------
    dict
        Mapping with ``units`` (list of ``{filename, stem, digest, csv_path}``)
        and ``failures`` (list of ``{filename, error}``), both sorted by
        filename.
    """
    src_dir = Path(src_dir)
    units = []
    failures = []
    for path in sorted(src_dir.glob("*.xlsx")):
        if path.name.startswith(_LOCK_PREFIX):
            continue
        try:
            unit = ingest_workbook(path)
            csv_path = write_unit(unit, out_dir)
        except Exception as error:  # noqa: BLE001 - record, never abort a sweep
            failures.append({"filename": path.name, "error": repr(error)})
            continue
        units.append(
            {
                "filename": path.name,
                "stem": unit.stem,
                "digest": unit.digest,
                "csv_path": os.fspath(csv_path),
            }
        )
    return {"units": units, "failures": failures}
