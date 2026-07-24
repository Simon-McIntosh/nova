"""Structural parser for coil-metrology workbooks.

A measured coil-alignment workbook is a loose grid, not a table: each sheet
(one assembly phase) carries one or two side-by-side coil blocks, and above
each block an optional rigid-body transform. This module locates that
structure in an :mod:`openpyxl` workbook and yields typed intermediate
objects, without deciding how the values are serialized.

Two passes over the same file are required. The first pass
(``data_only=False``) exposes cell formulas so that a value produced by a
spreadsheet formula (``=VLOOKUP(...)``) is flagged rather than mistaken for a
literal. The second pass (``data_only=True``) exposes the cached value that
the spreadsheet last computed -- that cached number is the measurement. A
formula cell whose cache is absent yields a ``None`` value and an anomaly, so
the measurement is recorded as missing rather than silently dropped.

The workbook is opened read-only; the source corpus is treated as immutable.
"""

from __future__ import annotations

from dataclasses import dataclass, field

import openpyxl

# Column offsets from a block's leftmost ("Coil") header cell. A block is
# either six columns wide (no uncertainty) or nine (with 2-sigma uncertainty).
_OFF_POINT = 1
_OFF_NAME = 2
_OFF_COORD = {"x": 3, "y": 4, "z": 5}
_OFF_UNC = {"x": 6, "y": 7, "z": 8}

WIDTH_NO_UNCERTAINTY = 6
WIDTH_WITH_UNCERTAINTY = 9

COORD_AXES = ("x", "y", "z")

# A transform block sits above a coil block sharing its coordinate columns.
# Its "transform" label is two columns left of the coordinate columns, i.e.
# in the block's Name column; the six components occupy the X..uZ columns.
_TRANSFORM_LABEL_OFFSET = 2
TRANSFORM_AXES = ("dx", "dy", "dz", "rx", "ry", "rz")
TRANSFORM_UNITS = {
    "dx": "mm",
    "dy": "mm",
    "dz": "mm",
    "rx": "deg",
    "ry": "deg",
    "rz": "deg",
}
# Component column offsets from the associated coil header column.
_OFF_TRANSFORM = {axis: 3 + index for index, axis in enumerate(TRANSFORM_AXES)}

_COIL_HEADER = "Coil"
_TRANSFORM_LABEL = "transform"
_METADATA_SHEET = "Metadata"


@dataclass(frozen=True)
class Cell:
    """A single numeric measurement cell and its provenance.

    Parameters
    ----------
    value : float or None
        The cached numeric value, or ``None`` when the cell is empty or a
        formula whose cache is absent.
    is_formula : bool
        ``True`` when the source cell held a spreadsheet formula.
    """

    value: float | None
    is_formula: bool


EMPTY_CELL = Cell(value=None, is_formula=False)


@dataclass
class Fiducial:
    """One measured point: a coordinate triple with optional uncertainty.

    Parameters
    ----------
    point_group_raw : str
        The point-group label as written in the sheet (before normalization).
    name_raw : str
        The fiducial name as written (before normalization).
    coords : dict of str to Cell
        Coordinate cells keyed by axis ``"x"``, ``"y"``, ``"z"``.
    uncerts : dict of str to Cell
        Uncertainty cells keyed by axis; :data:`EMPTY_CELL` when the block
        carries no uncertainty columns.
    """

    point_group_raw: str
    name_raw: str
    coords: dict[str, Cell]
    uncerts: dict[str, Cell]


@dataclass
class CoilBlock:
    """A coil's table of measured points on one sheet.

    Parameters
    ----------
    coil : int or None
        Coil number read from the ``Coil`` column, or ``None`` if absent.
    header_row, header_col : int
        One-based row and column of the ``Coil`` header cell.
    width : int
        Block width in columns, 6 or 9.
    has_uncertainty : bool
        ``True`` when the block carries ``uX``/``uY``/``uZ`` columns.
    fiducials : list of Fiducial
        Measured points, top to bottom.
    """

    coil: int | None
    header_row: int
    header_col: int
    width: int
    has_uncertainty: bool
    fiducials: list[Fiducial] = field(default_factory=list)


@dataclass
class Transform:
    """A rigid-body transform above a coil block.

    Parameters
    ----------
    coil : int or None
        Coil number of the associated block.
    label_row, label_col : int
        One-based position of the ``transform`` label cell.
    header_col : int
        Header column of the associated coil block.
    components : dict of str to Cell
        Component cells keyed by ``dx``, ``dy``, ``dz`` (mm) and ``rx``,
        ``ry``, ``rz`` (deg).
    populated : bool
        ``True`` when at least one component carries a value.
    """

    coil: int | None
    label_row: int
    label_col: int
    header_col: int
    components: dict[str, Cell]
    populated: bool


@dataclass
class SheetParse:
    """Parsed structure of one worksheet.

    Parameters
    ----------
    name : str
        Sheet title, verbatim.
    blocks : list of CoilBlock
        Coil blocks, ordered left to right by header column.
    transforms : list of Transform
        Transform blocks, ordered left to right.
    anomalies : list of str
        Structural anomalies noted while parsing this sheet.
    """

    name: str
    blocks: list[CoilBlock] = field(default_factory=list)
    transforms: list[Transform] = field(default_factory=list)
    anomalies: list[str] = field(default_factory=list)


@dataclass
class WorkbookParse:
    """Parsed structure of a whole workbook.

    Parameters
    ----------
    sheets : list of SheetParse
        Measurement sheets in workbook order (the ``Metadata`` sheet is
        excluded here; its text is captured separately).
    metadata_text : list of tuple
        ``(row, col, text)`` triples for non-empty text cells on the
        ``Metadata`` sheet, sorted by position.
    anomalies : list of str
        Workbook-level anomalies.
    """

    sheets: list[SheetParse] = field(default_factory=list)
    metadata_text: list[tuple[int, int, str]] = field(default_factory=list)
    anomalies: list[str] = field(default_factory=list)


def _coerce_number(value) -> tuple[float | None, bool]:
    """Coerce a raw cell value to a float.

    Parameters
    ----------
    value : object
        Raw cell value from openpyxl.

    Returns
    -------
    tuple
        ``(number, is_text)`` where ``number`` is the float value or ``None``
        and ``is_text`` flags a non-empty non-numeric cell in a numeric slot.
    """
    if value is None:
        return None, False
    if isinstance(value, bool):
        return float(value), False
    if isinstance(value, (int, float)):
        return float(value), False
    try:
        return float(value), False
    except (TypeError, ValueError):
        return None, True


def _grids(worksheet):
    """Return dense value and data-type maps for a worksheet.

    Parameters
    ----------
    worksheet : openpyxl worksheet
        Read-only worksheet to scan.

    Returns
    -------
    tuple
        ``(values, dtypes, max_row, max_col)`` where ``values`` and
        ``dtypes`` map ``(row, col)`` to the cell value and data type.
    """
    values: dict[tuple[int, int], object] = {}
    dtypes: dict[tuple[int, int], str] = {}
    max_row = 0
    max_col = 0
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            key = (cell.row, cell.column)
            values[key] = cell.value
            dtypes[key] = cell.data_type
            if cell.row > max_row:
                max_row = cell.row
            if cell.column > max_col:
                max_col = cell.column
    return values, dtypes, max_row, max_col


def _find_labels(values, label):
    """Return sorted positions of cells whose stripped text equals ``label``.

    Parameters
    ----------
    values : dict
        ``(row, col)`` to value map.
    label : str
        Target label.

    Returns
    -------
    list of tuple
        ``(row, col)`` positions, sorted.
    """
    hits = [
        key
        for key, value in values.items()
        if isinstance(value, str) and value.strip() == label
    ]
    return sorted(hits)


def _block_width(values, header_row, header_col):
    """Return the contiguous header width starting at a ``Coil`` cell."""
    width = 0
    col = header_col
    while (header_row, col) in values and isinstance(values[(header_row, col)], str):
        width += 1
        col += 1
    return width


def _read_cell(values, dtypes, formula_dtypes, row, col):
    """Build a :class:`Cell` from the data-only value and formula data type."""
    is_formula = formula_dtypes.get((row, col)) == "f"
    number, _is_text = _coerce_number(values.get((row, col)))
    return Cell(value=number, is_formula=is_formula), _is_text


def _read_block(
    values,
    dtypes,
    formula_dtypes,
    header_row,
    header_col,
    max_row,
    anomalies,
):
    """Read one coil block into a :class:`CoilBlock`.

    Parameters
    ----------
    values, dtypes : dict
        Data-only value and data-type maps.
    formula_dtypes : dict
        Data-type map from the formula pass (used to detect formulas).
    header_row, header_col : int
        Position of the ``Coil`` header cell.
    max_row : int
        Last populated row in the sheet.
    anomalies : list of str
        Sheet anomaly list, appended in place.

    Returns
    -------
    CoilBlock
        The parsed block.
    """
    width = _block_width(values, header_row, header_col)
    has_uncertainty = width >= WIDTH_WITH_UNCERTAINTY
    block = CoilBlock(
        coil=None,
        header_row=header_row,
        header_col=header_col,
        width=width,
        has_uncertainty=has_uncertainty,
    )

    block_cols = range(header_col, header_col + width)
    retained: list[tuple[int, dict]] = []
    pending_spacers: list[int] = []
    for row in range(header_row + 1, max_row + 1):
        present = [(row, col) in values for col in block_cols]
        if not any(present):
            pending_spacers.append(row)
            continue
        if retained and pending_spacers:
            for spacer_row in pending_spacers:
                anomalies.append(
                    f"spacer row {spacer_row} inside coil block at "
                    f"column {header_col} (header row {header_row})"
                )
        pending_spacers = []

        point_raw = values.get((row, header_col + _OFF_POINT))
        name_raw = values.get((row, header_col + _OFF_NAME))
        coil_value = values.get((row, header_col))
        if block.coil is None and coil_value is not None:
            number, is_text = _coerce_number(coil_value)
            if number is not None:
                block.coil = int(round(number))
            elif not is_text:
                block.coil = None

        coords = {}
        uncerts = {}
        for axis in COORD_AXES:
            cell, is_text = _read_cell(
                values,
                dtypes,
                formula_dtypes,
                row,
                header_col + _OFF_COORD[axis],
            )
            coords[axis] = cell
            if is_text:
                anomalies.append(
                    f"non-numeric value in coordinate column at row {row}, "
                    f"axis {axis}, block column {header_col}"
                )
            if has_uncertainty:
                ucell, u_is_text = _read_cell(
                    values,
                    dtypes,
                    formula_dtypes,
                    row,
                    header_col + _OFF_UNC[axis],
                )
                uncerts[axis] = ucell
                if u_is_text:
                    anomalies.append(
                        f"non-numeric value in uncertainty column at row "
                        f"{row}, axis {axis}, block column {header_col}"
                    )
            else:
                uncerts[axis] = EMPTY_CELL

        formula_axes = [
            axis
            for axis in COORD_AXES
            if coords[axis].is_formula and coords[axis].value is None
        ]
        formula_axes += [
            f"u{axis}"
            for axis in COORD_AXES
            if uncerts[axis].is_formula and uncerts[axis].value is None
        ]
        if formula_axes:
            anomalies.append(
                f"formula cell without cached value at row {row}, block "
                f"column {header_col}, axes {','.join(formula_axes)}"
            )

        retained.append(
            (
                row,
                {
                    "point_raw": point_raw,
                    "name_raw": name_raw,
                    "coords": coords,
                    "uncerts": uncerts,
                },
            )
        )

    # Forward-fill the point-group label across rows that omit it.
    last_group = ""
    for _row, entry in retained:
        raw = entry["point_raw"]
        if raw is not None and str(raw).strip() != "":
            last_group = str(raw)
        name = entry["name_raw"]
        block.fiducials.append(
            Fiducial(
                point_group_raw=last_group,
                name_raw="" if name is None else str(name),
                coords=entry["coords"],
                uncerts=entry["uncerts"],
            )
        )
    return block


def _match_header_col(block_cols, label_col):
    """Return the coil header column associated with a transform label.

    Parameters
    ----------
    block_cols : list of int
        Header columns of the sheet's coil blocks.
    label_col : int
        Column of the ``transform`` label.

    Returns
    -------
    int
        The associated coil header column: exact match at
        ``label_col - 2`` when present, otherwise the nearest block column.
    """
    target = label_col - _TRANSFORM_LABEL_OFFSET
    if target in block_cols:
        return target
    if not block_cols:
        return target
    return min(block_cols, key=lambda col: abs(col - target))


def _read_transform(
    values, dtypes, formula_dtypes, label_row, label_col, header_col, coil
):
    """Read a transform block into a :class:`Transform`."""
    components = {}
    populated = False
    value_row = label_row + 1
    for axis in TRANSFORM_AXES:
        cell, _is_text = _read_cell(
            values,
            dtypes,
            formula_dtypes,
            value_row,
            header_col + _OFF_TRANSFORM[axis],
        )
        components[axis] = cell
        if cell.value is not None or cell.is_formula:
            populated = True
    return Transform(
        coil=coil,
        label_row=label_row,
        label_col=label_col,
        header_col=header_col,
        components=components,
        populated=populated,
    )


def _parse_sheet(name, values, dtypes, formula_dtypes, max_row):
    """Parse a single worksheet into a :class:`SheetParse`."""
    sheet = SheetParse(name=name)
    header_positions = _find_labels(values, _COIL_HEADER)
    header_positions.sort(key=lambda pos: pos[1])
    for header_row, header_col in header_positions:
        sheet.blocks.append(
            _read_block(
                values,
                dtypes,
                formula_dtypes,
                header_row,
                header_col,
                max_row,
                sheet.anomalies,
            )
        )

    block_cols = [block.header_col for block in sheet.blocks]
    coil_by_col = {block.header_col: block.coil for block in sheet.blocks}
    transform_positions = _find_labels(values, _TRANSFORM_LABEL)
    transform_positions.sort(key=lambda pos: pos[1])
    for label_row, label_col in transform_positions:
        header_col = _match_header_col(block_cols, label_col)
        sheet.transforms.append(
            _read_transform(
                values,
                dtypes,
                formula_dtypes,
                label_row,
                label_col,
                header_col,
                coil_by_col.get(header_col),
            )
        )
    return sheet


def _metadata_text(values):
    """Return sorted ``(row, col, text)`` triples for non-empty text cells."""
    entries = [
        (row, col, str(value))
        for (row, col), value in values.items()
        if isinstance(value, str) and value.strip() != ""
    ]
    entries.sort()
    return entries


def parse_workbook(path) -> WorkbookParse:
    """Parse a coil-metrology workbook into typed intermediate objects.

    Parameters
    ----------
    path : os.PathLike or str
        Workbook to read. Opened read-only.

    Returns
    -------
    WorkbookParse
        Parsed sheets, ``Metadata`` text, and anomalies.
    """
    formula_book = openpyxl.load_workbook(path, data_only=False, read_only=True)
    data_book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        parse = WorkbookParse()
        for worksheet in data_book.worksheets:
            name = worksheet.title
            values, dtypes, max_row, _max_col = _grids(worksheet)
            if name == _METADATA_SHEET:
                parse.metadata_text = _metadata_text(values)
                continue
            formula_values, formula_dtypes, _fr, _fc = _grids(formula_book[name])
            parse.sheets.append(
                _parse_sheet(name, values, dtypes, formula_dtypes, max_row)
            )
        return parse
    finally:
        formula_book.close()
        data_book.close()
