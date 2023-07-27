"""Manage TFC fiducial data for coil and sector allignment."""
from contextlib import contextmanager
from dataclasses import dataclass, field
from functools import cached_property
from glob import glob
import os
from pathlib import Path
import pickle
from typing import ClassVar

import numpy as np
import openpyxl
import pandas

from nova.assembly.fiducialccl import Fiducial, FiducialRE, FiducialIDM
from nova.database.filepath import FilePath


@dataclass
class SectorFile:
    r"""
    Sector filename path and group.

    Sector module data is downloaded from IDM and stored in IO shared folder at:
    \\\\io-ws-ccstore1\\ANSYS_Data\\mcintos\\sector_modules

    data_dir : str
        Data directory. Set as mount point location to access IO shared folder
    """

    sector: int
    filename: str = ""
    version: str | int = "latest"
    data_dir: str = "/mnt/share/sector_modules"

    def __post_init__(self):
        """Locate source datafiles."""
        self.locate_file()

    def locate_file(self):
        """Locate remote datafile and set version and set version number."""
        match self.filename:
            case "":
                self._locate_file()
            case str():
                self._update_metadata()

    def _update_metadata(self):
        """Update sector and version number from filename."""
        self.version = self._get_version(self.filename)
        self.sector = self._get_sector(self.filename)

    @staticmethod
    def _get_version(filename: str) -> float:
        """Return filename version."""
        return float(filename.split("v")[-1].split(".")[0].replace("_", "."))

    @staticmethod
    def _get_sector(filename: str) -> int:
        """Return sector index."""
        return int(filename.split("#")[-1].split("_")[0])

    @cached_property
    def filenames(self) -> list[str]:
        """Return filename list."""
        return [
            Path(path).with_suffix("").name
            for path in glob(self.data_dir + f"/*Sector_Module_#{self.sector}*.xlsx")
        ]

    @cached_property
    def versions(self) -> list[float]:
        """Return version list."""
        versions = [self._get_version(filename) for filename in self.filenames]
        if len(versions) == 0:
            raise FileNotFoundError(
                f"Unable to locate RE data files at {self.data_dir}. "
                "Check contents of data directory. "
                "Confirm connection to the IO network if data_dir is a mounted share."
            )
        return versions

    def _locate_file(self):
        """Locate source datafile and update filename and version."""
        match self.version:
            case "latest":
                index = np.argmax(self.versions)
            case int() | float():
                index = self.versions.index(self.version)
            case _:
                raise ValueError(
                    f"version {self.version} not latest or in {self.versions}"
                )
        self.filename = self.filenames[index]
        self.version = self.versions[index]


@dataclass
class SectorData(FilePath, SectorFile):
    """Manage fiducial coil and sector assembly data sourced from IDM."""

    dirname: Path | str = ".nova/sector_modules"
    data: dict = field(init=False, repr=False, default_factory=dict)
    ccl: dict = field(init=False, repr=False, default_factory=dict)

    def __post_init__(self):
        """Load / build dataset."""
        super().__post_init__()
        self.load_build()

    def load_build(self):
        """Load or build dataset."""
        try:
            self.load()
        except (FileNotFoundError, OSError):
            self.build()

    def load(self):
        """Load pickled dataset."""
        with open(self.filepath.with_suffix(".pickle"), "rb") as file:
            self.data = pickle.load(file)
            self.ccl = pickle.load(file)

    def store(self):
        """Pickle data and ccl."""
        with open(self.filepath.with_suffix(".pickle"), "wb") as file:
            pickle.dump(self.data, file, protocol=5)
            pickle.dump(self.ccl, file, protocol=5)

    def build(self):
        """Build mesurment dataset."""
        self.build_data()
        self.build_ccl()
        self.store()

    @property
    def xls_file(self):
        """Return xls filename."""
        return os.path.join(self.data_dir, f"{self.filename}.xlsx")

    def _initialize_data(self):
        """Initialize data as a bare nested dict with coil name entries."""
        self.data = {name: {} for name in self._coil_names("Nominal")}

    def build_data(self):
        """Build dataset."""
        with self.openbook():
            self._initialize_data()
            for worksheet in self.book:
                sheet = worksheet.title
                if sheet == "Metadata":
                    continue
                for index, name in enumerate(self._coil_names(sheet)):
                    self.data[name][sheet] = self.read_frame(index, sheet)

    @cached_property
    def coil(self) -> list[str]:
        """Return list of coil names."""
        return [name for name in self.data]

    @cached_property
    def phase(self) -> list[str]:
        """Return list of assembly phases."""
        return [phase for phase in self.data[self.coil[0]] if phase != "Nominal"]

    def _initalize_ccl(self):
        """Init ccl as bare nested dict with assembly stage entries."""
        self.ccl = {phase: {coil: None for coil in self.coil} for phase in self.phase}

    def build_ccl(self):
        """Build ccl data."""
        self._initalize_ccl()
        for coil in self.coil:
            nominal = self.data[coil]["Nominal"]
            for phase in self.phase:
                self.ccl[phase][coil] = self.data[coil][phase].loc[nominal.index]
                try:
                    self.ccl[phase][coil].loc[:, nominal.columns] -= nominal
                except TypeError:
                    self.ccl[phase][coil].loc[:, nominal.columns] = 0.0
                self.ccl[phase][coil] = self.ccl[phase][coil].rename(
                    columns={col: f"d{col}" for col in "xyz"}
                )
                self.ccl[phase][coil] = self.ccl[phase][coil].rename(index={"F'": "F"})
                self.ccl[phase][coil].index = self.ccl[phase][coil].index.droplevel(
                    [0, 1]
                )
                self.ccl[phase][coil].index.name = None

    @contextmanager
    def openbook(self):
        """Manage access to source workbook."""
        self.book = openpyxl.load_workbook(self.xls_file, data_only=True)
        yield
        self.book.close()

    def locate(self, item: str, sheet: str):
        """Return item row/column locations in worksheet."""
        index = []
        for col in self.book[sheet].iter_cols():
            for cell in col:
                if cell.value == item:
                    index.append((cell.row, cell.column))
                    break
        assert len(index) == 2
        return index

    def _coil_index(self, sheet: str):
        """Return list dataset origins."""
        return self.locate("Coil", sheet)

    def _coil_names(self, sheet: str):
        """Return list of coil names."""
        name = []
        for row, cell in self._coil_index(sheet):
            name.append(self.book[sheet].cell(row + 1, cell).value)
        return name

    def _column_number(self, index, sheet: str):
        """Return column number."""
        for ncol, cell in enumerate(
            self.book[sheet].iter_cols(
                min_row=index[0], max_row=index[0], min_col=index[1]
            )
        ):
            if cell[0].value is None:
                ncol -= 1
                break
        return ncol + 1

    def read_frame(self, coil: int, sheet: str):
        """Return pandas dataframe from indexed sheet."""
        index = self._coil_index(sheet)[coil]
        ncol = self._column_number(index, sheet)
        usecols = list(range(index[1] - 1, index[1] - 1 + ncol))
        data = pandas.read_excel(
            self.xls_file,
            sheet_name=sheet,
            skiprows=index[0] - 1,
            usecols=usecols,
            index_col=[0, 1, 2],
            keep_default_na=False,
        )
        data = data.rename(
            columns={col: col.split(".")[0].lower() for col in data.columns}
        )
        data.index.rename(
            [name.split(".")[0] for name in data.index.names], inplace=True
        )
        return data


@dataclass
class FiducialSector(Fiducial):
    """Manage Reverse Engineering fiducial data."""

    phase: str = "FAT supplier"
    sectors: list[int] = field(default_factory=lambda: [*range(1, 9)])
    variance: dict[str, pandas.DataFrame] | dict = field(
        init=False, default_factory=dict
    )

    sheets: ClassVar[dict[str, str]] = {"FATsup": "FAT supplier", "SSAT": "SSAT BR"}

    def __post_init__(self):
        """Propogate origin."""
        self._set_phase()
        super().__post_init__()
        self.source = "Reverse Engineering IDM datasets (xls workbooks)"
        self.origin = [self.origin[coil - 1] for coil in self.delta]
        self._load_variance()

    def _set_phase(self):
        """Expand short string phase label."""
        self.phase = self.sheets.get(self.phase, self.phase)

    def _load_deltas(self):
        """Implement load deltas abstractmethod."""
        columns = ["dx", "dy", "dz"]
        for sector in self.sectors:
            data = SectorData(sector)
            for coil, ccl in data.ccl[self.phase].items():
                self.delta[coil] = ccl.loc[self.target, columns]

    def _load_variance(self):
        columns = ["ux", "uy", "uz"]
        for sector in self.sectors:
            data = SectorData(sector)
            for coil, ccl in data.ccl[self.phase].items():
                two_sigma = ccl.loc[self.target, columns]
                self.variance[coil] = (two_sigma / 2) ** 2
                self.variance[coil] = self.variance[coil].rename(
                    columns={col: f"s2{col[-1]}" for col in columns}
                )

    def compare(self, source="RE"):
        """Compare fiducial sector data with previous RE dataset."""
        match source:
            case "IDM":
                previous = FiducialIDM()
            case "RE":
                previous = FiducialRE()
            case _:
                raise ValueError(f"source {source} not in [RE, IDM]")

        for coil, ccl in self.delta.items():
            if coil not in previous.delta:
                continue
            _ccl = previous.delta[coil]
            if source == "RE":
                _ccl = _ccl = previous.delta[coil].xs("FAT", 1)
            change = ccl.loc[:, ["dx", "dy", "dz"]] - _ccl
            if not np.allclose(np.array(change, float), 0):
                print(f"\ncoil #{coil}")
                print(change)


if __name__ == "__main__":
    # sector = SectorData(8)

    fiducial = FiducialSector(phase="FATsup")  # , sectors=[8]
    fiducial.compare("RE")
    # fiducial.plot()

    # for coil, ccl in fiducial.delta.items():
    #    print(f"Coil {coil}")
    #    print(ccl)
    #    print()
