"""Manage access to IDM TFC metrology data."""

from contextlib import contextmanager
from dataclasses import dataclass, field
from functools import cached_property
import os
from pathlib import Path
import pickle

import numpy as np
import openpyxl
import pandas

from nova.assembly.sectorfile import SectorFile
from nova.database.filepath import FilePath


@dataclass
class SectorData(FilePath, SectorFile):
    """Manage fiducial coil and sector assembly data sourced from IDM."""

    dirname: Path | str = ".nova/sector_modules"
    data: dict = field(init=False, repr=False, default_factory=dict)
    ccl: dict = field(init=False, repr=False, default_factory=dict)
    coil: list = field(init=False, default_factory=list)

    def __post_init__(self):
        """Load / build dataset."""
        super().__post_init__()
        self.load_build()

    def load_build(self):
        """Load or build dataset."""
        try:
            self.load()
        except (FileNotFoundError, OSError, EOFError):
            self.build()

    def load(self):
        """Load pickled dataset."""
        with open(self.filepath.with_suffix(".pickle"), "rb") as file:
            self.data = pickle.load(file)
            self.ccl = pickle.load(file)
            self.coil = pickle.load(file)

    def store(self):
        """Pickle data and ccl."""
        with open(self.filepath.with_suffix(".pickle"), "wb") as file:
            pickle.dump(self.data, file, protocol=5)
            pickle.dump(self.ccl, file, protocol=5)
            pickle.dump(self.coil, file, protocol=5)

    def build(self):
        """Build mesurment dataset."""
        self.build_data()
        self.build_ccl()
        self.store()

    @property
    def xls_file(self):
        """Return xls filename."""
        return os.path.join(self.datadir, f"{self.filename}.xlsx")

    @property
    def _xls_file(self):
        """Return saveas xls filename."""
        return os.path.join(self.datadir, f"_{self.filename}.xlsx")

    def _initialize_data(self):
        """Initialize data as a bare nested dict with coil name entries."""
        self.data = {name: {} for name in self._coil_names("Nominal")}

    def build_data(self):
        """Build dataset."""
        with self.openbook():
            self._initialize_data()
            for worksheet in self.book:
                sheet = worksheet.title
                if sheet == "Metadata":
                    continue
                for index, name in enumerate(self._coil_names(sheet)):
                    self.data[name][sheet] = self.read_frame(index, sheet)
            self.coil = self._coil_names("Nominal")
        for coil in self.coil:
            if (
                self.data[coil]["Nominal"].values.dtype == np.object_
                or np.isnan(self.data[coil]["Nominal"].values).any()
            ):
                self.data.pop(coil)

    @contextmanager
    def savebook(self):
        """Save workbook to _xlsfile."""
        yield
        self.book.save(self._xls_file)

    def write(self, worksheet, xls_index, data):
        """Append data to workbook."""
        for i in range(data.shape[0]):
            for j in range(data.shape[1]):
                worksheet.cell(i + xls_index[0] + 1, j + xls_index[1] + 3, data[i, j])

    @cached_property
    def phase(self) -> list[str]:
        """Return list of assembly phases."""
        return [phase for phase in self.data[self.coil[0]] if phase != "Nominal"]

    def _initalize_ccl(self):
        """Init ccl as bare nested dict with assembly stage entries."""
        self.ccl = {phase: {coil: None for coil in self.coil} for phase in self.phase}

    def build_ccl(self):
        """Build ccl data."""
        self._initalize_ccl()
        for coil in self.coil:
            if coil not in self.data:
                for phase in self.phase:
                    self.ccl[phase].pop(coil)
                continue
            nominal = self.data[coil]["Nominal"]
            for phase in self.phase:
                self.ccl[phase][coil] = self.data[coil][phase].loc[nominal.index]
                try:
                    self.ccl[phase][coil].loc[:, nominal.columns] -= nominal
                except TypeError:
                    print("type error", coil, phase)
                    self.ccl[phase][coil].loc[:, nominal.columns] = 0.0
                self.ccl[phase][coil] = self.ccl[phase][coil].rename(
                    columns={col: f"d{col}" for col in "xyz"}
                )
                self.ccl[phase][coil] = self.ccl[phase][coil].rename(index={"F'": "F"})
                self.ccl[phase][coil].index = self.ccl[phase][coil].index.droplevel(
                    [0, 1]
                )
                self.ccl[phase][coil].index.name = None

    @contextmanager
    def openbook(self):
        """Manage access to source workbook."""
        self.book = openpyxl.load_workbook(self.xls_file, data_only=False)
        yield
        self.book.close()

    def locate(self, item: str, sheet: str):
        """Return item row/column locations in worksheet."""
        index = []
        for col in self.book[sheet].iter_cols():
            for cell in col:
                if cell.value == item:
                    index.append((cell.row, cell.column))
                    break
        assert len(index) == 2
        return index

    def _coil_index(self, sheet: str):
        """Return list of dataset origins."""
        return self.locate("Coil", sheet)

    def _coil_names(self, sheet: str):
        """Return list of coil names."""
        name = []
        for row, cell in self._coil_index(sheet):
            name.append(self.book[sheet].cell(row + 1, cell).value)
        return name

    def _column_number(self, index, sheet: str):
        """Return column number."""
        for ncol, cell in enumerate(
            self.book[sheet].iter_cols(
                min_row=index[0], max_row=index[0], min_col=index[1]
            )
        ):
            if cell[0].value is None:
                ncol -= 1
                break
        return ncol + 1

    def read_frame(self, coil: int, sheet: str):
        """Return pandas dataframe from indexed sheet."""
        index = self._coil_index(sheet)[coil]
        ncol = self._column_number(index, sheet)
        usecols = list(range(index[1] - 1, index[1] - 1 + ncol))
        data = pandas.read_excel(
            self.xls_file,
            sheet_name=sheet,
            skiprows=index[0] - 1,
            usecols=usecols,
            index_col=[0, 1, 2],
            keep_default_na=False,
            na_values="",
            dtype=dict.fromkeys(["X", "Y", "Z", "uX", "uY", "uZ"], float),
        )
        data = data.rename(
            columns={col: col.split(".")[0].lower() for col in data.columns}
        )
        data.index.rename(
            [name.split(".")[0] for name in data.index.names], inplace=True
        )
        return data


if __name__ == "__main__":
    sector = SectorData(7)
    # sector.build()
