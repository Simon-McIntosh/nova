"""Manage access to non-axisymmetric coil data."""

from dataclasses import dataclass, field
from functools import cached_property
from typing import ClassVar
import warnings

import numpy as np
import openpyxl
from packaging import version
from tqdm import tqdm

from nova.biot.biotframe import Source
from nova.biot.space import Segment
from nova.geometry.polygeom import Polygon
from nova.geometry.polyline import PolyLine
from nova.geometry.section import Section
from nova.graphics.plot import Plot
from nova.imas.coil import coil_name, coil_names, part_name
from nova.imas.dataset import ImasIds
from nova.imas.machine import CoilDatabase
from nova.imas.scenario import Scenario


@dataclass
class Elements(Plot):
    """Manage conductor element data extracted from coils_non_axisymmetric IDS."""

    elements: ImasIds = field(repr=False, default=None)
    points: np.ndarray = field(
        init=False, repr=False, default_factory=lambda: np.array([])
    )
    polyline: PolyLine = field(init=False, repr=False, default_factory=PolyLine)
    data: dict[str, np.ndarray] = field(init=False, repr=False, default_factory=dict)

    point_attrs: ClassVar[list[str]] = [
        "start_points",
        "intermediate_points",
        "end_points",
    ]

    def __post_init__(self):
        """Extract point data from elements node."""
        self.data["types"] = self.elements.types
        for attr in self.point_attrs:
            self.data[attr] = self._to_array(getattr(self.elements, attr))
        self.points = self._extract_points()
        self.polyline = self._extract_polyline()
        self.polyline.points = self.points

    def __len__(self):
        """Return maximum length of arrays stored in data dict."""
        return np.max([len(array) for array in self.data.values()])

    def _to_array(self, ids_points, attrs=["r", "phi", "z"]):
        """Return cartesian point array from nested cylindrical ids structure."""
        data = {
            label: getattr(ids_points, attr)
            for label, attr in zip(["r", "phi", "z"], attrs)
        }
        return np.c_[
            data["r"] * np.cos(data["phi"]),
            data["r"] * np.sin(data["phi"]),
            data["z"],
        ]

    @cached_property
    def _point_array(self):
        """Return shape (n, 3, 3) point array."""
        points = np.zeros((len(self), 3, 3))
        for i, attr in enumerate(self.point_attrs):
            if len(self.data[attr]) == 0:  # skip unfilled entries
                continue
            points[..., i] = self.data[attr]
        return points

    @cached_property
    def _type_array(self):
        """Return shape (n,) element types array."""
        types = np.zeros(len(self))
        types[:] = self.data["types"]
        return types

    def _extract_points(self):
        """Return shape (n, 3) unique point centerline."""
        points = []
        for element_type, point_array in zip(self._type_array, self._point_array):
            match element_type:
                case 1:  # line
                    points.append(point_array[..., 0])
                case 2:  # arc
                    points.append(point_array[..., 0])
                    points.append(point_array[..., 1])
                case _:
                    raise ValueError(f"element_type {element_type} not supported")
        points.append(point_array[..., -1])
        return np.array(points)[
            np.sort(np.unique(points, return_index=True, axis=0)[1])
        ]

    def _extract_polyline(self):
        """Return segmented polyline."""
        polyline = PolyLine(minimum_arc_nodes=3, filament=True)
        for element_type, point_array in zip(self._type_array, self._point_array):
            match element_type:
                case 1:  # line
                    normal = point_array[..., 1] - point_array[..., 0]
                    points = np.stack(
                        [point_array[..., 0], point_array[..., 2]], axis=0
                    )
                    polyline.append(points, normal[np.newaxis, :])
                case 2:  # arc
                    points = np.stack(
                        [point_array[..., 0], point_array[..., 1], point_array[..., 2]],
                        axis=0,
                    )
                    polyline.append(points)
                case _:
                    raise NotImplementedError(
                        f"element type {element_type} not implemented"
                    )
        return polyline

    @property
    def start_axes(self):
        """Return start axes."""
        start_segment = self.polyline.segments[0]
        return Segment(
            start_segment.normal, start_segment.axis, start_segment.name
        ).start_axes

    def plot(self):
        """Plot polyline."""
        self.set_axes("3d")
        self.axes.plot(*self.points.T)


@dataclass
class CoilsNonAxisymmetric(Plot, CoilDatabase, Scenario):
    """Manage access to coils_non_axisymmetric ids."""

    pulse: int = 115001
    run: int = 1
    name: str = "coils_non_axisymmetric"
    ids_node: str = "coil"

    coil_attrs: ClassVar[list[str]] = ["turns", "resistance"]

    def build(self):
        """Build netCDF database using data extracted from imasdb."""
        name = coil_names(self.ids.coil)
        with self.build_scenario():
            self.data.coords["coil_name"] = name
            self.data.coords["coil_index"] = "coil_name", range(len(name))
            self.data.coords["point"] = ["x", "y", "z"]

            points = {}
            for coil in tqdm(self.ids.coil, "building coils non-axisymmetric"):
                name = coil_name(coil).value
                points[name] = []
                for i, conductor in enumerate(coil.conductor):
                    elements = Elements(
                        elements=conductor.elements,
                    )
                    points[name].extend(elements.points)
                    if self.ids_dd_version <= version.Version("3.39"):  # IDS version
                        try:
                            section = Section(
                                elements._to_array(
                                    conductor.cross_section,
                                    attrs=["delta_r", "delta_phi", "delta_z"],
                                ),
                                triad=elements.start_axes,
                            )
                            section.to_axes(np.identity(3))
                            polygon = Polygon(section.points[:, 1:])
                        except AttributeError as error:
                            warnings.warn(
                                "cross section structure unreachable for "
                                f"ids written with dd {self.ids_dd_version} < 3.39. "
                                f"{error.__str__()}"
                            )
                            polygon = Polygon({"c": [0, 0, 0.05]})
                    else:
                        cross_section = conductor.cross_section[0]
                        match index := cross_section.geometry_type.index:
                            case 2:
                                assert cross_section.geometry_type.name == "circular"
                                width = cross_section.width
                                polygon = Polygon({"circle": [0, 0, width, width, 2]})
                            case 4:
                                assert cross_section.geometry_type.name == "square"
                                width = cross_section.width
                                polygon = Polygon({"square": [0, 0, width]})
                            case 5:
                                assert cross_section.geometry_type.name == "annulus"
                                width = cross_section.width
                                radius_inner = cross_section.radius_inner
                                thickness = 1 - radius_inner / (width / 2)
                                polygon = Polygon({"pipe": [0, 0, width, thickness, 2]})
                            case _:
                                raise NotImplementedError(
                                    f"Geometry type index {index} not implemented."
                                )

                    if i > 0:
                        name = f"name{i}"
                    elements.polyline.cross_section = polygon.points
                    self.winding.insert(
                        polyline=elements.polyline,
                        cross_section=polygon,
                        name=name,
                        part=part_name(coil),
                        delim="_",
                    )
                    if i > 0:
                        self.linkframe([coil_name(coil), name])
                points[name] = np.array(points[name])
            maximum_point_number = np.max([len(_points) for _points in points.values()])
            self.data.coords["point_index"] = np.arange(1, maximum_point_number + 1)
            self.data["points"] = ("coil_name", "point_index", "point"), np.zeros(
                (
                    self.data.sizes["coil_name"],
                    self.data.sizes["point_index"],
                    self.data.sizes["point"],
                )
            )
            self.data["points_length"] = "coil_name", [
                len(_points) for _points in points.values()
            ]
            for i, (number, name) in enumerate(
                zip(self.data["points_length"].data, self.data["coil_name"].data)
            ):
                self.data["points"].data[i, :number] = points[name]

    def _write_header(self, worksheet):
        """Write worksheet header."""
        for col, coord in enumerate("XYZ"):
            worksheet.cell(1, col + 3, f"{coord} Coord")

    def _write_data(self, worksheet, xls_index, data):
        """Append data to workbook."""
        self._write_header(worksheet)
        for i in range(data.shape[0]):
            for j in range(data.shape[1]):
                worksheet.cell(i + xls_index[0], j + xls_index[1], data[i, j])

    def write_excel(self):
        """Write centerline path to an excel file [mm]."""
        workbook = openpyxl.Workbook()
        for i, coil in enumerate(self.frame.index):
            worksheet = workbook.create_sheet(coil, i)
            index = self.subframe.index[self.subframe.frame == coil]
            source = Source(self.subframe.loc[index, :].to_dict())
            self._write_data(worksheet, (2, 3), 1e3 * source.space.path)
        workbook.save(self.filepath.with_suffix(".xlsx"))
        workbook.close()


if __name__ == "__main__":
    # cc_ids = CoilsNonAxisymmetric(111003, 2)  # CC
    # cs_ids = CoilsNonAxisymmetric(111004, 1)  # CS

    # elm_ids = CoilsNonAxisymmetric(115001, 2)  # ELM

    ids = CoilsNonAxisymmetric(111006, 1)

    ids.plot()

    # coil = elm_ids  # + cc_ids  # + cs_ids
    # coil.plot()
    # coil.frame.vtkplot()
    # coil3d._clear()
