"""Builders for synthetic coil-metrology workbooks.

These helpers construct workbooks in memory with :mod:`openpyxl`, laid out
exactly as the measured corpus is: side-by-side coil blocks headed
``Coil|Point|Name|X|Y|Z[|uX|uY|uZ]``, with optional rigid-body transform
blocks above them. They let the transcoder tests exercise every structural
variant the corpus contains without touching the read-only source data.

A formula cell is written by giving its coordinate a string beginning with
``=``. Because no spreadsheet engine runs, such a cell has no cached value --
which is exactly the corpus case of a formula whose cache is absent.
"""

from __future__ import annotations

import openpyxl

# Column offsets within a block, mirroring the corpus layout.
_OFF = {
    "Coil": 0,
    "Point": 1,
    "Name": 2,
    "X": 3,
    "Y": 4,
    "Z": 5,
    "uX": 6,
    "uY": 7,
    "uZ": 8,
}
_HEADERS_6 = ["Coil", "Point", "Name", "X", "Y", "Z"]
_HEADERS_9 = _HEADERS_6 + ["uX", "uY", "uZ"]
_TRANSFORM_LABELS = ["dx [mm]", "dy [mm]", "dz [mm]", "X [deg]", "Y [deg]", "Z [deg]"]


def write_block(
    worksheet,
    *,
    header_row,
    header_col,
    coil,
    points,
    has_uncertainty=True,
):
    """Write one coil block into a worksheet.

    Parameters
    ----------
    worksheet : openpyxl worksheet
        Target worksheet.
    header_row, header_col : int
        One-based position of the ``Coil`` header cell.
    coil : int
        Coil number written once, at the first data row.
    points : list of dict
        One entry per measured point with keys ``group`` (str or ``None`` to
        forward-fill), ``name`` (str), ``xyz`` (three-tuple of floats,
        strings, or ``None``), and optionally ``u`` (three-tuple) for
        uncertainty. A ``None`` entry writes an all-empty spacer row.
    has_uncertainty : bool
        Write the nine-column header when ``True``, else six columns.
    """
    headers = _HEADERS_9 if has_uncertainty else _HEADERS_6
    for name in headers:
        worksheet.cell(header_row, header_col + _OFF[name], name)

    row = header_row + 1
    written_coil = False
    for entry in points:
        if entry is None:
            row += 1
            continue
        if not written_coil:
            worksheet.cell(row, header_col + _OFF["Coil"], coil)
            written_coil = True
        if entry.get("group") is not None:
            worksheet.cell(row, header_col + _OFF["Point"], entry["group"])
        worksheet.cell(row, header_col + _OFF["Name"], entry["name"])
        for axis, key in zip("xyz", ("X", "Y", "Z")):
            value = entry["xyz"]["xyz".index(axis)]
            if value is not None:
                worksheet.cell(row, header_col + _OFF[key], value)
        if has_uncertainty and entry.get("u") is not None:
            for axis, key in zip("xyz", ("uX", "uY", "uZ")):
                value = entry["u"]["xyz".index(axis)]
                if value is not None:
                    worksheet.cell(row, header_col + _OFF[key], value)
        row += 1


def write_transform(worksheet, *, header_col, label_row, components):
    """Write a rigid-body transform block above a coil block.

    Parameters
    ----------
    worksheet : openpyxl worksheet
        Target worksheet.
    header_col : int
        Header column of the associated coil block.
    label_row : int
        Row of the ``transform`` label; values go on ``label_row + 1``.
    components : list or None
        Six component values ``[dx, dy, dz, rx, ry, rz]`` (``None`` entries
        left blank), or ``None`` to write only the labels (empty transform).
    """
    worksheet.cell(label_row, header_col + 2, "transform")
    for index, label in enumerate(_TRANSFORM_LABELS):
        worksheet.cell(label_row, header_col + 3 + index, label)
    if components is None:
        return
    for index, value in enumerate(components):
        if value is not None:
            worksheet.cell(label_row + 1, header_col + 3 + index, value)


def build_workbook(sheets, *, metadata_text=None):
    """Build a workbook from a declarative sheet specification.

    Parameters
    ----------
    sheets : list of tuple
        ``(name, build)`` pairs where ``build`` is a callable taking the
        worksheet and populating it.
    metadata_text : list of tuple, optional
        ``(row, col, text)`` triples written to a ``Metadata`` sheet.

    Returns
    -------
    openpyxl.Workbook
        The populated workbook.
    """
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    if metadata_text is not None:
        sheet = workbook.create_sheet("Metadata")
        for row, col, text in metadata_text:
            sheet.cell(row, col, text)
    for name, build in sheets:
        build(workbook.create_sheet(name))
    return workbook
