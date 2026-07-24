"""Unit tests for the fiducial-fit core (nova.assembly.fiducialfit).

Pure-function tests run everywhere; the fit-driver tests rebuild sector
workbooks from the in-repo canonical units via the characterization fixture
machinery and skip visibly when those fixtures cannot be materialized.
"""

import shutil

import matplotlib

matplotlib.use("Agg", force=True)

import numpy as np
import pandas as pd
import pytest
from scipy.spatial.transform import Rotation

from nova.assembly.fiducialfit import FiducialFit, fit_sector
from tests.characterization import _fixtures

FIXTURES_AVAILABLE = _fixtures.sector_modules_available()

requires_fixtures = pytest.mark.skipif(
    not FIXTURES_AVAILABLE,
    reason="sector-module fixtures cannot be materialized: in-repo canonical "
    "units or the workbook transcoder are unavailable",
)


class TestErrorVector:
    """error_vector reduces per-axis deltas over the constraint index sets."""

    def setup_method(self):
        rng = np.random.default_rng(7)
        self.delta = rng.normal(size=(8, 3))

    def test_rms(self):
        error = FiducialFit.error_vector(self.delta, "rms")
        index = FiducialFit.fiducial_index
        assert error[0] == pytest.approx(np.mean(self.delta[index["radial"], 0] ** 2))
        assert error[1] == pytest.approx(np.mean(self.delta[index["toroidal"], 1] ** 2))
        assert error[2] == pytest.approx(np.mean(self.delta[index["vertical"], 2] ** 2))

    def test_max(self):
        error = FiducialFit.error_vector(self.delta, "max")
        index = FiducialFit.fiducial_index
        assert error[0] == pytest.approx(np.max(abs(self.delta[index["radial"], 0])))
        assert error[1] == pytest.approx(np.max(abs(self.delta[index["toroidal"], 1])))
        assert error[2] == pytest.approx(np.max(abs(self.delta[index["vertical"], 2])))

    def test_unknown_method_raises(self):
        with pytest.raises(NotImplementedError):
            FiducialFit.error_vector(self.delta, "median")


class TestTransform:
    """transform applies a translation then an intrinsic-Euler rotation."""

    def test_translation_only(self):
        fit = FiducialFit.__new__(FiducialFit)
        points = np.zeros((4, 3))
        moved = fit.transform(np.array([1.0, -2.0, 3.0]), points)
        assert np.allclose(moved, [1.0, -2.0, 3.0])

    def test_six_dof(self):
        fit = FiducialFit.__new__(FiducialFit)
        rng = np.random.default_rng(11)
        points = rng.normal(size=(5, 3))
        x = np.array([0.5, -0.25, 1.0, 5.0, -3.0, 45.0])
        moved = fit.transform(x, points.copy())
        rotate = Rotation.from_euler("XYZ", x[-3:], degrees=True)
        assert np.allclose(moved, rotate.apply(points + x[:3]))

    def test_input_not_mutated(self):
        fit = FiducialFit.__new__(FiducialFit)
        points = np.ones((3, 3))
        fit.transform(np.array([1.0, 1.0, 1.0]), points)
        assert np.allclose(points, 1.0)


class TestJoin:
    def test_with_postfix(self):
        assert FiducialFit.join("error", "gpr") == "error_gpr"

    def test_without_postfix(self):
        assert FiducialFit.join("error", "") == "error"


class TestMaskFrame:
    """mask_frame blanks axes of fiducials outside the controlled index sets."""

    def test_masks_uncontrolled_deltas(self):
        names = list("ABCDEFGH")
        frame = pd.DataFrame(
            np.ones((len(names), 3)), index=names, columns=["x", "y", "z"]
        )
        masked = FiducialFit.mask_frame(frame)
        index = FiducialFit.fiducial_index
        for axis, key in zip("xyz", index):
            controlled = set(np.array(names)[index[key]])
            for name in names:
                value = masked.loc[name, axis]
                if name in controlled:
                    assert value == 1.0
                else:
                    assert value == ""


@pytest.fixture(scope="module")
def fitted():
    """Fit the sector-7 coil pair from the rebuilt SSAT BR workbook."""
    if not FIXTURES_AVAILABLE:
        pytest.skip(
            "sector-module fixtures cannot be materialized: in-repo canonical "
            "units or the workbook transcoder are unavailable"
        )
    _fixtures.ensure_sector_cache()
    return fit_sector(
        phase="SSAT BR",
        sectors={7: [8, 9]},
        fill=False,
        infer=True,
        ilis=True,
        ilis_pcr=True,
        method="rms",
        coupled=False,
        private=False,
    )


@requires_fixtures
class TestFitSector:
    def test_transform_solution_shape(self, fitted):
        assert fitted.data.opt_x.dims == ("coil", "transform")
        assert fitted.data.opt_x.shape == (2, 6)
        assert np.all(np.isfinite(fitted.data.opt_x.data))

    def test_fit_reduces_objective(self, fitted):
        """The optimized transform beats the identity on the fitted point set."""
        zero = np.zeros(fitted.data.sizes["transform"])
        for coil in fitted.data.coil.values:
            points = fitted.points(coil)
            opt_x = fitted.data.opt_x.sel(coil=coil).data
            assert fitted.scalar_error(opt_x, points) <= fitted.scalar_error(
                zero, points
            )

    def test_delta_summary(self, fitted):
        for target in ["fiducial_fit", "fiducial_fit_gpr"]:
            summary = fitted.delta_summary(target)
            assert isinstance(summary, pd.DataFrame)
            coils = {level for level, _ in summary.columns}
            assert coils == {"Coil 8", "Coil 9"}

    def test_write_round_trip(self, fitted, tmp_path, monkeypatch):
        """write() lands the fitted transform in the target workbook sheet."""
        import openpyxl

        from nova.assembly.sectordata import SectorData
        from nova.assembly.sectorfile import SectorFile

        for path in _fixtures.WORKBOOK_CACHE.glob("*.xlsx"):
            shutil.copy2(path, tmp_path / path.name)
        pickle_dir = tmp_path / "pickles"
        pickle_dir.mkdir()
        for cls in (SectorData, SectorFile):
            defaults = list(cls.__init__.__defaults__)
            defaults[_fixtures._DATADIR_INDEX] = str(tmp_path)
            if cls is SectorData:
                defaults[_fixtures._DIRNAME_INDEX] = pickle_dir
            monkeypatch.setattr(cls.__init__, "__defaults__", tuple(defaults))

        sheet = "SSAT test target"
        fitted.write(sheet)

        # savebook lands in the private (leading underscore) sibling workbook
        workbook_path = next(tmp_path.glob("_Sector_Module_#7_*_v15_0.xlsx"))
        book = openpyxl.load_workbook(workbook_path)
        assert sheet in book.sheetnames
        cells = [
            cell.value
            for row in book[sheet].iter_rows()
            for cell in row
            if cell.value is not None
        ]
        assert "transform" in cells
        for value in fitted.data.opt_x.sel(coil=8).data:
            assert any(
                isinstance(cell, float) and cell == pytest.approx(value)
                for cell in cells
            )
